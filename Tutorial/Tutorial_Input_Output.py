__author__ = 'admin'

'''_______________________________ csv ____________________________ '''
import csv
### deal with row by bow
with open(inputpath, 'rb') as fileread, open(outputpath, 'wb') as filewrite:
    reader = csv.reader(fileread)
    writer = csv.writer(filewrite)
    next(reader, None)  # skip the headers

    for row_read in reader:

        symbol = row_read[0]+exchange_symbol
        name = get_name(symbol).replace('"','')
        writer.writerow([symbol, name])
        print symbol, name
    filewrite.close()
    print "Done"


'''_______________________ NO module _____________________ '''
with open('symbolslist.txt') as ifp, open('intermediateresults.csv', 'a') as results_fp:
    # do not use csv.reader/writer
    for row in ifp:
        row = row.strip()
        api = [ get_price(row), get_market_cap(row) ]
        api = re.sub("\[\'|\'|\]", "", str(api))
        results_fp.write(str(row) +"," +str(api) +"\n")     # not .writerow
        print row, api

symbols = open("symbolslist.txt")
readsymbols = symbols.read()




'''______________________ pandas ______________________________ '''

import pandas as pd
security_map = pd.read_csv(security_map_path)
Dataframe.to_csv(root + 'Exchange1.csv')

template = pd.read_excel(root + "non_us_signal_copy1.xlsx")
df.to_excel('path.xlxs')


writer = pd.ExcelWriter('output.xlsx')
writer.date_format = None # <--- Workaround for date formatting
writer.datetime_format = None  # <--- this one for datetime
df1.to_excel(writer,'Sheet1')
df2.to_excel(writer,'Sheet2')


'''_________________________ xlrd _______________'''
import xlrd

template_book = xlrd.open_workbook(template_path)
template_sheet = template_book.sheet_by_index(0)


'''_______________ openpyxl ________________ '''
import openpyxl
wb = openpyxl.load_workbook('C:\Users\mglobe10\Desktop\Price & Volume Databases\Price & Volume config.xlsx')
ws = wb.get_sheet_by_name(name='Europe Intraday Data')
resttime =ws.cell('C9').value

wb2 = openpyxl.Workbook()  # create new excel file
sheeti = wb2.create_sheet(title = newname)
wb2.remove_sheet(blank_sheet)
sheeti.cell(row=1,column=1).value = 'Date'
wb2.save(newpathfile)
number_row = ws.get_highest_row()

ws.append([1, 2, 3])


'''_______________________ xlsxwriter ____________________________________ '''
import xlsxwriter

if not os.path.isfile(params['Output File Path']+'\\'+params['Output File Name'] + '.xlsx'):
    workbook = xlsxwriter.Workbook('Daily Performance Metrics.xlsx')
    worksheet_1 = workbook.add_worksheet(params['Trading Account 1'] + 'Account')
    merge_format = workbook.add_format({'bold': 1,
                                        'border': 1,
                                        'align': 'center',
                                        'valign': 'vcenter'})

    worksheet_1.merge_range('B1:C1', 'NAV Details', merge_format)











