import openpyxl
import pandas as pd
import datetime as dt


root = 'D:\\poly\\courses\\book\\python\\tutorial\\data\\'

series1 = pd.Series([dt.datetime.strptime('20040304', '%Y%m%d').date(),
                   1234567.1234,
                   'short',
                   0.3432])
series2 = pd.Series([dt.datetime.strptime('20130304', '%Y%m%d').date(),
                   7654321.1234,
                   'short',
                   0.03422])
df = pd.DataFrame([series1,series2])


def write_df(df, wb, ws_name):
    ws = wb.get_sheet_by_name(name = ws_name)
    high_row = ws.get_highest_row()
    for i in range(len(df)):
        for j in range(4):
            font_fix = ws.cell(row = high_row, column = j+1).font
            border_fix = ws.cell(row = high_row, column = j+1).border
            fill_fix = ws.cell(row = high_row, column = j+1).fill
            almt_fix = ws.cell(row = high_row, column = j+1).alignment
            num_fix = ws.cell(row = high_row, column = j+1).number_format

            c = ws.cell(row = high_row + 1 + i, column = j+1)
            c.value = df.ix[i,j]
            c.font = font_fix
            c.border = border_fix
            c.fill = fill_fix
            c.alignment = almt_fix
            c.number_format = num_fix



# wb = openpyxl.load_workbook(root + 'openpyxl.xlsx')
write_df(df,wb,'mysheet')




ws_1 = wb.get_sheet_by_name(name = 'Sheet1')
ws_2 = wb.get_sheet_by_name(name = 'mysheet')
ws_sum = wb.get_sheet_by_name(name = 'summary')
high_row_1 = ws_1.get_hightest_row()
high_row_sum = ws_sum.get_hightest_row()

for i in range(len(df)/2):
    for j in range(4):
        font_fix = ws_sum.cell(row = high_row_sum, column = j+1).font
        border_fix = ws_sum.cell(row = high_row_sum, column = j+1).border
        fill_fix = ws_sum.cell(row = high_row_sum, column = j+1).fill
        almt_fix = ws_sum.cell(row = high_row_sum, column = j+1).alignment
        num_fix = ws_sum.cell(row = high_row_sum, column = j+1).number_format

        c = ws_sum.cell(row = high_row_sum + 1 + i, column = j+1)

        if j == 2:
            if ws_sum.cell(row = high_row_sum + 1 + i, column = 3) > 0:
                c.value = 'Long'
            else:
                c.value = 'Short'
        else:
            c.value = ws_1.cell(row = high_row_sum + 1 + i, column = j+1).value + ws_2.cell(row = high_row_sum + 1 + i, column = j+1).value

        c.font = font_fix
        c.border = border_fix
        c.fill = fill_fix
        c.alignment = almt_fix
        c.number_format = num_fix


wb.save(root + 'openpyxl.xlsx')



